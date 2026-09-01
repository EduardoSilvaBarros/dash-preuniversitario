"""
===============================================================================
DASHBOARD EJECUTIVO ESTRATÉGICO — LANZAMIENTO PREUNIVERSITARIO SISE
===============================================================================
Aplicación Interactiva en Streamlit para Análisis C-Level (Brandbook SISE 2022):
  - Carga Ultra-rápida desde colegios_sise_5km.csv (491 KB en lugar de DBF pesado)
  - Identidad Visual SISE Completa: Sidebar en Negro SISE #231F20, Accent Rojo #FF0E49
  - Leyenda del Sidebar en Cuadro Blanco (#FFFFFF) de Alto Contraste
  - Gráfico Comparativo Dual: Alumnos Objetivo Meta SISE (#FF0E49) vs Mercado Preu Asignado (#0075B0)
  - Modelo Macro de Mercado: Tabla de Proyección y Embudo de Conversión (Funnel Chart)
  - Filtros Multiselección y Rendimiento Canvas a 60 FPS
"""

import os, math, json, struct, subprocess
import pandas as pd
import openpyxl
import streamlit as st
import streamlit.components.v1 as components
import plotly.express as px
import plotly.graph_objects as go

# =============================================================================
# CONFIGURACIÓN PÁGINA STREAMLIT
# =============================================================================
st.set_page_config(
    page_title="SISE » Dashboard Ejecutivo Preuniversitario",
    page_icon="🎓",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Estilos CSS Oficiales Brandbook SISE
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Barlow+Condensed:ital,wght@0,400;0,600;0,700;0,900;1,400&display=swap');
    
    :root {
        --sise-red:     #FF0E49;
        --sise-black:   #231F20;
        --sise-white:   #FFFFFF;
        --sise-surface: #F8F9FA;
        --sise-teal:    #00B2A9;
        --sise-green:   #009860;
        --sise-navy:    #0075B0;
        --sise-sky:     #5BC6E8;
        --sise-coral:   #EA2839;
        --sise-yellow:  #F9E11E;
        --sise-purple:  #952D98;
    }
    
    html, body, [class*="css"] {
        font-family: 'Barlow Condensed', sans-serif !important;
        color: var(--sise-black);
    }
    
    /* STYLING DEL SIDEBAR COMPLETO CON BRANDING SISE (#231F20) */
    [data-testid="stSidebar"] {
        background-color: #231F20 !important;
        color: #FFFFFF !important;
        border-right: 2px solid var(--sise-red);
    }
    [data-testid="stSidebar"] h2, [data-testid="stSidebar"] h3 {
        color: var(--sise-red) !important;
        font-weight: 900 !important;
        text-transform: uppercase !important;
        letter-spacing: 0.5px;
    }
    [data-testid="stSidebar"] p, [data-testid="stSidebar"] label, [data-testid="stSidebar"] span {
        color: #FFFFFF !important;
    }
    [data-testid="stSidebar"] .stMultiSelect [data-baseweb="tag"] {
        background-color: var(--sise-red) !important;
        color: #FFFFFF !important;
        border-radius: 3px !important;
    }
    
    /* Header Flecha Contenedor SISE */
    .sise-header-container {
        background: var(--sise-red);
        color: #FFFFFF;
        padding: 20px 28px;
        border-radius: 6px;
        position: relative;
        overflow: hidden;
        margin-bottom: 20px;
        box-shadow: 0 4px 12px rgba(255,14,73,0.25);
    }
    .sise-header-container::after {
        content: '';
        position: absolute;
        right: -40px; top: 0;
        width: 180px; height: 100%;
        background: var(--sise-black);
        clip-path: polygon(30% 0, 100% 0, 100% 100%, 0% 100%);
    }
    .sise-header-title {
        font-weight: 900;
        font-size: 32px;
        text-transform: uppercase;
        letter-spacing: -0.01em;
        margin: 0;
        line-height: 1;
    }
    .sise-header-subtitle {
        font-weight: 600;
        font-size: 15px;
        opacity: 0.95;
        margin-top: 6px;
        text-transform: uppercase;
        letter-spacing: 0.5px;
    }
    .sise-tag {
        display: inline-block;
        background: var(--sise-black);
        color: #FFFFFF;
        font-weight: 900;
        font-size: 11px;
        padding: 3px 8px;
        border-radius: 3px;
        margin-top: 8px;
        letter-spacing: 1px;
    }

    /* KPI Cards SISE (#FF0E49) */
    .kpi-card {
        background: #FFFFFF;
        padding: 14px;
        border-radius: 4px;
        border-left: 4px solid var(--sise-red);
        box-shadow: 0 2px 6px rgba(0,0,0,0.06);
    }
    .kpi-title {
        font-size: 12px;
        font-weight: 700;
        color: var(--sise-black);
        opacity: 0.75;
        text-transform: uppercase;
        letter-spacing: 0.5px;
    }
    .kpi-value {
        font-size: 26px;
        font-weight: 900;
        color: var(--sise-red);
        margin-top: 2px;
        line-height: 1;
    }
    .kpi-sub {
        font-size: 11px;
        color: var(--sise-green);
        font-weight: 700;
        margin-top: 3px;
    }
    
    /* Reseñas y Acordeones SISE */
    .brand-summary-card { background: #FFFFFF; padding: 12px 15px; border-left: 4px solid var(--sise-navy); border-radius: 4px; margin-bottom: 10px; font-size: 13px; color: var(--sise-black); line-height: 1.35; box-shadow: 0 1px 3px rgba(0,0,0,0.05); }
    .review-quote-box { background: #FFFFFF; padding: 8px 12px; border-left: 3px solid var(--sise-red); border-radius: 4px; margin-bottom: 6px; font-size: 12px; color: var(--sise-black); font-style: italic; line-height: 1.35; box-shadow: 0 1px 3px rgba(0,0,0,0.04); }
    .sede-meta-line { font-size: 12px; color: var(--sise-black); margin-bottom: 6px; padding: 5px 10px; background: #EAECEE; border-radius: 4px; font-weight: 600; }
    
    /* CUADRO BLANCO PARA LA LEYENDA EN EL SIDEBAR (#FFFFFF) */
    .sidebar-legend {
        background: #FFFFFF !important;
        padding: 14px;
        border-radius: 6px;
        border: 1px solid #D0D3D4;
        font-size: 11px;
        margin-top: 15px;
        box-shadow: 0 4px 8px rgba(0,0,0,0.15);
    }
    .sidebar-legend-title {
        font-weight: 900;
        color: var(--sise-red) !important;
        margin-bottom: 8px;
        text-transform: uppercase;
        font-size: 12px;
        letter-spacing: 0.5px;
    }
    .sidebar-legend-item {
        display: flex;
        align-items: center;
        gap: 8px;
        margin-bottom: 5px;
        font-weight: 700;
        color: #231F20 !important;
    }
    .badge-dot { width: 11px; height: 11px; border-radius: 50%; display: inline-block; flex-shrink: 0; }

    /* Pestañas (Tabs) Estilizadas */
    button[data-baseweb="tab"] {
        font-weight: 700 !important;
        font-size: 14px !important;
        text-transform: uppercase !important;
    }
    button[aria-selected="true"] {
        color: var(--sise-red) !important;
        border-bottom-color: var(--sise-red) !important;
    }
    
    /* Footer SISE */
    .sise-footer { text-align: center; padding: 15px; color: var(--sise-black); font-size: 12px; font-weight: 700; border-top: 2px solid var(--sise-red); margin-top: 30px; }
</style>
""", unsafe_allow_html=True)

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PRE_PROCESSED_CSV = os.path.join(SCRIPT_DIR, 'colegios_sise_5km.csv')
DBF_FILE = os.path.abspath(os.path.join(SCRIPT_DIR, '..', 'ESCALE', 'Padron_web_20260805', 'Padron_web.dbf'))
COMPETENCIA_FILE = os.path.join(SCRIPT_DIR, 'ubicaciones_academias_lima.xlsx')
DEMANDA_FILE = os.path.join(SCRIPT_DIR, 'Demanda_Preuniversitario_SISE (5).xlsx')

RADIO_KM = 5.0

SISE_CHART_PALETTE = ['#FF0E49', '#00B2A9', '#0075B0', '#F9E11E', '#009860', '#952D98', '#5BC6E8', '#EA2839']

SISE_CAMPUSES = {
    "SISE ATE": {"lat": -12.019876204903364, "lon": -76.90082170490433, "distrito": "Ate", "pob_2025": 617099, "mercado_asignado": 7176, "alumnos_obj": 359, "short": "Ate"},
    "SISE CANTUARIAS": {"lat": -12.121194637015256, "lon": -77.02566624663862, "distrito": "Miraflores", "pob_2025": 69434, "mercado_asignado": 807, "alumnos_obj": 40, "short": "Cantuarias"},
    "SISE COMAS": {"lat": -12.011699284781184, "lon": -77.08076911835867, "distrito": "Comas", "pob_2025": 633491, "mercado_asignado": 7366, "alumnos_obj": 368, "short": "Comas"},
    "SISE INDEPENDENCIA": {"lat": -11.990264141393677, "lon": -77.05937114110782, "distrito": "Independencia", "pob_2025": 235922, "mercado_asignado": 2743, "alumnos_obj": 137, "short": "Independencia"},
    "SISE PUENTE PIEDRA": {"lat": -11.864461775095888, "lon": -77.07785528252775, "distrito": "Puente Piedra", "pob_2025": 263002, "mercado_asignado": 3058, "alumnos_obj": 153, "short": "Puente Piedra"},
    "SISE SAN JUAN DE LURIGANCHO": {"lat": -12.017285992993552, "lon": -77.0043344380135, "distrito": "San Juan de Lurigancho", "pob_2025": 1319232, "mercado_asignado": 15340, "alumnos_obj": 767, "short": "SJL"},
    "SISE SAN JUAN DE MIRAFLORES": {"lat": -12.160543157379358, "lon": -76.97334683598721, "distrito": "San Juan de Miraflores", "pob_2025": 539197, "mercado_asignado": 6270, "alumnos_obj": 313, "short": "SJM"},
    "SISE SAN MIGUEL": {"lat": -12.079667293957085, "lon": -77.07728819755285, "distrito": "San Miguel", "pob_2025": 155021, "mercado_asignado": 1803, "alumnos_obj": 90, "short": "San Miguel"},
    "SISE SANTA BEATRIZ": {"lat": -12.077679313532318, "lon": -77.03590747486689, "distrito": "Cercado de Lima", "pob_2025": 308548, "mercado_asignado": 3588, "alumnos_obj": 179, "short": "Santa Beatriz"},
    "SISE SURCO": {"lat": -12.146043182168036, "lon": -76.98792208996211, "distrito": "Santiago de Surco", "pob_2025": 326580, "mercado_asignado": 3797, "alumnos_obj": 190, "short": "Surco"},
    "SISE VENTANILLA": {"lat": -11.86960227457678, "lon": -77.12853022023926, "distrito": "Ventanilla", "pob_2025": 271948, "mercado_asignado": 3162, "alumnos_obj": 174, "short": "Ventanilla"},
    "SISE VILLA EL SALVADOR": {"lat": -12.191768167571478, "lon": -76.93744404914503, "distrito": "Villa El Salvador", "pob_2025": 506664, "mercado_asignado": 5891, "alumnos_obj": 295, "short": "VES"}
}

BRAND_EXECUTIVE_SUMMARIES = {
    "Aula 20": "📌 **Aula 20**: Cobertura en Lima Norte y Este. Reseñas destacan preparación para UNMSM/Agraria, con observaciones sobre espacio en aulas y alta intensidad comercial.",
    "Trilce": "📌 **Trilce**: Cadena tradicional consolidada. Destacan nivel en simulacros y exigencia docente; comentarios mencionan costo de mensualidad.",
    "Pamer": "📌 **Pamer**: Sistema de tutoría personalizada e intensiva. Usuarios valoran el seguimiento continuo al estudiante.",
    "César Vallejo y ADUNI": "📌 **César Vallejo / ADUNI**: Líder para UNI y UNMSM. Alta valoración académica; algunas opiniones mencionan alta densidad de alumnos por aula.",
    "CEPREPUCP": "📌 **CEPREPUCP**: Centro oficial de la PUCP con ingreso directo. Alta reputación en metodología e infraestructura.",
    "Pre San Marcos": "📌 **Pre San Marcos**: Centro oficial de la UNMSM. Muy solicitada por vacantes directas; feedback resalta exigencia.",
    "CEPRE UNI": "📌 **CEPRE UNI**: Centro oficial de la UNI. Nivel técnico exigente en ciencias exactas."
}

def haversine(lat1, lon1, lat2, lon2):
    R = 6371.0
    dlat = math.radians(lat2 - lat1)
    dlon = math.radians(lon2 - lon1)
    a = math.sin(dlat / 2)**2 + math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) * math.sin(dlon / 2)**2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
    return R * c

# =============================================================================
# CACHE DE DATOS CON STREAMLIT (LIGERO Y OPTIMIZADO PARA STREAMLIT CLOUD)
# =============================================================================
@st.cache_data
def load_all_data():
    df_demanda_sede = pd.DataFrame()
    df_mercado_macro = pd.DataFrame()
    
    # 1. Cargar Demanda & Mercado Macro
    temp_dem = os.path.join(SCRIPT_DIR, "_temp_dem_st.xlsx")
    if os.path.exists(DEMANDA_FILE):
        subprocess.run(f'powershell -Command "Copy-Item \'{DEMANDA_FILE}\' \'{temp_dem}\' -Force"', shell=True)
        read_dem = temp_dem if os.path.exists(temp_dem) else DEMANDA_FILE
        try:
            wb_dem = openpyxl.load_workbook(read_dem, data_only=True)
            
            if 'Mercado' in wb_dem.sheetnames:
                ws_m = wb_dem['Mercado']
                m_rows = []
                for r in [8, 9, 10, 11]:
                    u_name = str(ws_m.cell(r, 1).value or '').strip()
                    if u_name:
                        m_rows.append({
                            "Universidad": u_name,
                            "Postulantes 2025": float(ws_m.cell(r, 2).value or 0),
                            "Ingresantes 2025": float(ws_m.cell(r, 3).value or 0),
                            "% No Ingresa (info)": float(ws_m.cell(r, 4).value or 0),
                            "Postulantes 2026 (proy.)": float(ws_m.cell(r, 5).value or 0),
                            "% Alcanzable (Lima)": float(ws_m.cell(r, 6).value or 0) if ws_m.cell(r, 6).value else None,
                            "% Usa Academia": float(ws_m.cell(r, 7).value or 0) if ws_m.cell(r, 7).value else None,
                            "Mercado de Preu": float(ws_m.cell(r, 8).value or 0),
                            "% Market Share": float(ws_m.cell(r, 9).value or 0) if ws_m.cell(r, 9).value else None,
                            "Alumnos Objetivo": float(ws_m.cell(r, 10).value or 0)
                        })
                df_mercado_macro = pd.DataFrame(m_rows)

            if 'Demanda_Sede' in wb_dem.sheetnames:
                ws_ds = wb_dem['Demanda_Sede']
                dem_rows = []
                for r in range(5, ws_ds.max_row+1):
                    s_name = str(ws_ds.cell(r, 1).value or '').strip()
                    dist = str(ws_ds.cell(r, 2).value or '').strip()
                    if s_name and s_name != 'TOTAL':
                        try:
                            pob = float(ws_ds.cell(r, 3).value or 0)
                            peso = float(ws_ds.cell(r, 4).value or 0)
                            mercado = float(ws_ds.cell(r, 5).value or 0)
                            capt = float(ws_ds.cell(r, 6).value or 0)
                            obj = float(ws_ds.cell(r, 7).value or 0)
                            dem_rows.append({
                                "Sede": s_name, "Distrito": dist, "Poblacion_2025": pob,
                                "Peso_Distrital": peso, "Mercado_Preu_Asignado": mercado,
                                "Pct_Captacion": capt, "Alumnos_Objetivo": obj
                            })
                        except ValueError:
                            pass
                df_demanda_sede = pd.DataFrame(dem_rows)
            
            if os.path.exists(temp_dem):
                try: os.remove(temp_dem)
                except: pass
        except Exception:
            pass

    # 2. Cargar Competencia
    temp_comp = os.path.join(SCRIPT_DIR, "_temp_comp_st.xlsx")
    if os.path.exists(COMPETENCIA_FILE):
        subprocess.run(f'powershell -Command "Copy-Item \'{COMPETENCIA_FILE}\' \'{temp_comp}\' -Force"', shell=True)
        read_path = temp_comp if os.path.exists(temp_comp) else COMPETENCIA_FILE
        wb = openpyxl.load_workbook(read_path, data_only=True)
        ws = wb.active
        comp_data = []
        for r in range(2, ws.max_row+1):
            vals = [str(ws.cell(r, c).value if ws.cell(r, c).value is not None else '').strip() for c in range(1, 10)]
            if any(vals) and len(vals) >= 5:
                try:
                    lat, lon = float(vals[3]), float(vals[4])
                    comp_data.append({
                        "cadena": vals[0], "nombre": vals[1], "direccion": vals[2],
                        "latitud": lat, "longitud": lon, "calificacion": vals[5],
                        "total_resenas": vals[6], "resumen_resenas": vals[7] if len(vals)>7 else "",
                        "enlace_gmaps": vals[8] if len(vals)>8 else ""
                    })
                except ValueError:
                    pass
        df_comp = pd.DataFrame(comp_data)
        if os.path.exists(temp_comp):
            try: os.remove(temp_comp)
            except: pass
    else:
        df_comp = pd.DataFrame()

    if not df_comp.empty:
        comp_records = []
        for _, r in df_comp.iterrows():
            min_d = 9999.0
            close_c = None
            sedes_5k = []
            for cname, cinfo in SISE_CAMPUSES.items():
                d = haversine(cinfo['lat'], cinfo['lon'], r['latitud'], r['longitud'])
                if d < min_d:
                    min_d = d
                    close_c = cname
                if d <= RADIO_KM:
                    sedes_5k.append(cname)
            rec = r.to_dict()
            rec['sede_sise_cercana'] = close_c
            rec['distancia_sise_km'] = round(min_d, 2)
            rec['sedes_5km'] = sedes_5k
            rec['es_menor_5km'] = (min_d <= RADIO_KM)
            comp_records.append(rec)
        df_comp = pd.DataFrame(comp_records)

    # 3. Cargar Colegios (Prioriza colegios_sise_5km.csv ultra-ligero de 491 KB)
    if os.path.exists(PRE_PROCESSED_CSV):
        df_schools = pd.read_csv(PRE_PROCESSED_CSV, encoding='utf-8-sig')
    elif os.path.exists(DBF_FILE):
        schools_records = []
        with open(DBF_FILE, 'rb') as f:
            header = f.read(32)
            num_records, header_len, record_len = struct.unpack('<IHH', header[4:12])
            fields = []
            while True:
                b = f.read(1)
                if b == b'\x0d' or not b: break
                field_data = b + f.read(31)
                name = field_data[:11].rstrip(b'\x00').decode('latin1', errors='ignore')
                flen = field_data[16]
                fields.append((name, flen))
            f.seek(header_len)
            offsets = {}
            curr = 1
            for name, flen in fields:
                offsets[name] = (curr, flen)
                curr += flen

            dpto_o, dpto_l = offsets['D_DPTO']
            prov_o, prov_l = offsets['D_PROV']
            dist_o, dist_l = offsets['D_DIST']
            niv_o, niv_l = offsets['D_NIV_MOD']
            est_o, est_l = offsets['D_ESTADO']
            gest_o, gest_l = offsets['D_GESTION']
            name_o, name_l = offsets['CEN_EDU']
            dir_o, dir_l = offsets['DIR_CEN']
            lat_o, lat_l = offsets['NLAT_IE']
            lon_o, lon_l = offsets['NLONG_IE']
            cod_o, cod_l = offsets['COD_MOD']

            for _ in range(num_records):
                rec = f.read(record_len)
                if not rec or len(rec) < record_len or rec[0:1] == b'*': continue
                dpto = rec[dpto_o : dpto_o+dpto_l].decode('latin1', errors='ignore').strip().upper()
                if 'LIMA' in dpto or 'CALLAO' in dpto:
                    niv = rec[niv_o : niv_o+niv_l].decode('latin1', errors='ignore').strip().upper()
                    if 'SECUNDARIA' in niv:
                        est = rec[est_o : est_o+est_l].decode('latin1', errors='ignore').strip().upper()
                        if est != 'ACTIVO' and est != 'ACTIVA': continue
                        try:
                            slat = float(rec[lat_o : lat_o+lat_l].decode('latin1', errors='ignore').strip())
                            slon = float(rec[lon_o : lon_o+lon_l].decode('latin1', errors='ignore').strip())
                        except ValueError: continue
                        if slat == 0 or slon == 0: continue

                        prov = rec[prov_o : prov_o+prov_l].decode('latin1', errors='ignore').strip()
                        dist = rec[dist_o : dist_o+dist_l].decode('latin1', errors='ignore').strip()
                        gest = rec[gest_o : gest_o+gest_l].decode('latin1', errors='ignore').strip()
                        sname = rec[name_o : name_o+name_l].decode('latin1', errors='ignore').strip()
                        sdir = rec[dir_o : dir_o+dir_l].decode('latin1', errors='ignore').strip()
                        codmod = rec[cod_o : cod_o+cod_l].decode('latin1', errors='ignore').strip()

                        is_pub = any(k in gest.upper() for k in ['PÚBLICO', 'PUBLICO', 'ESTATAL', 'DIRECTA', 'CONVENIO'])

                        for cname, cinfo in SISE_CAMPUSES.items():
                            d_km = haversine(cinfo['lat'], cinfo['lon'], slat, slon)
                            if d_km <= RADIO_KM:
                                schools_records.append({
                                    "sede_sise": cname, "distancia_km": round(d_km, 2),
                                    "cod_mod": codmod, "nombre_colegio": sname, "gestion": gest,
                                    "tipo_gestion": "Público" if is_pub else "Privado",
                                    "provincia": prov, "distrito": dist, "direccion": sdir,
                                    "latitud": slat, "longitud": slon
                                })
        df_schools = pd.DataFrame(schools_records)
    else:
        df_schools = pd.DataFrame()

    return df_schools, df_comp, df_demanda_sede, df_mercado_macro

df_schools, df_comp, df_demanda_sede, df_mercado_macro = load_all_data()

# =============================================================================
# SIDEBAR / FILTROS CON BRANDING SISE (#231F20 Y #FF0E49)
# =============================================================================
st.sidebar.markdown("## 🎛️ FILTROS DE CONTROL")

sedes_disponibles = list(SISE_CAMPUSES.keys())
sedes_seleccionadas = st.sidebar.multiselect(
    "» 1. Seleccionar Sedes SISE:",
    options=sedes_disponibles,
    default=sedes_disponibles,
    help="Selecciona una o varias sedes para compararlas."
)

gestiones_seleccionadas = st.sidebar.multiselect(
    "» 2. Gestión de Colegios:",
    options=["Privado", "Público"],
    default=["Privado", "Público"]
)

cadenas_disponibles = sorted(list(df_comp['cadena'].unique())) if not df_comp.empty else []
cadenas_seleccionadas = st.sidebar.multiselect(
    "» 3. Marcas de Competencia:",
    options=cadenas_disponibles,
    default=cadenas_disponibles
)

# CUADRO BLANCO PARA LA LEYENDA (#FFFFFF) DE ALTO CONTRASTE REQUERIDO POR EL USUARIO
st.sidebar.markdown("""
<div class="sidebar-legend">
    <div class="sidebar-legend-title">🎨 Leyenda de Marcadores & Colores</div>
    <div class="sidebar-legend-item"><span class="badge-dot" style="background:#FF0E49; border: 2px solid #231F20;"></span> <strong>Sedes SISE (5 km)</strong></div>
    <hr style="border:0; border-top:1px solid #E0E0E0; margin:4px 0;">
    <div class="sidebar-legend-item"><span class="badge-dot" style="background:#00a8ff;"></span> Aula 20 (Azul Claro)</div>
    <div class="sidebar-legend-item"><span class="badge-dot" style="background:#e67e22;"></span> Trilce (Naranja)</div>
    <div class="sidebar-legend-item"><span class="badge-dot" style="background:#ffffff; border:1px solid #231F20;"></span> Pamer (Blanco)</div>
    <div class="sidebar-legend-item"><span class="badge-dot" style="background:#8e44ad;"></span> César Vallejo / ADUNI (Morado)</div>
    <div class="sidebar-legend-item"><span class="badge-dot" style="background:#0c2461;"></span> CEPREPUCP (Azul Oscuro)</div>
    <div class="sidebar-legend-item"><span class="badge-dot" style="background:#800020;"></span> Pre San Marcos (Guinda)</div>
    <div class="sidebar-legend-item"><span class="badge-dot" style="background:#f9e79f; border:1px solid #7d6608;"></span> CEPRE UNI (Crema)</div>
    <hr style="border:0; border-top:1px solid #E0E0E0; margin:4px 0;">
    <div class="sidebar-legend-item"><span class="badge-dot" style="background:#009860;"></span> Colegio Público (Verde 5px)</div>
    <div class="sidebar-legend-item"><span class="badge-dot" style="background:#0075B0;"></span> Colegio Privado (Azul 5px)</div>
</div>
""", unsafe_allow_html=True)

# =============================================================================
# ENCABEZADO CON FIRMA VISUAL "FLECHA CONTENEDOR" SISE (#FF0E49)
# =============================================================================
st.markdown("""
<header class="sise-header-container">
  <div class="sise-header-title">SISE » INSTITUTO SUPERIOR</div>
  <div class="sise-header-subtitle">ESTRATEGIA DE LANZAMIENTO PREUNIVERSITARIO — DASHBOARD EJECUTIVO C-LEVEL</div>
  <span class="sise-tag">#HAZLABIEN</span>
</header>
""", unsafe_allow_html=True)

# Aplicar Filtros Multiselección
df_sch_filtered = df_schools[
    df_schools['sede_sise'].isin(sedes_seleccionadas) &
    df_schools['tipo_gestion'].isin(gestiones_seleccionadas)
].copy() if not df_schools.empty else pd.DataFrame()

if not df_comp.empty:
    df_comp_filtered = df_comp[
        df_comp['cadena'].isin(cadenas_seleccionadas) &
        df_comp['sedes_5km'].apply(lambda sedes: any(s in sedes_seleccionadas for s in sedes))
    ].copy()
else:
    df_comp_filtered = pd.DataFrame()

# =============================================================================
# BLOQUE 1: TARJETAS KPI EJECUTIVAS
# =============================================================================
col1, col2, col3, col4, col5 = st.columns(5)

tot_colegios = len(df_sch_filtered)
tot_pub = len(df_sch_filtered[df_sch_filtered['tipo_gestion'] == 'Público']) if not df_sch_filtered.empty else 0
tot_priv = len(df_sch_filtered[df_sch_filtered['tipo_gestion'] == 'Privado']) if not df_sch_filtered.empty else 0
pct_priv = (tot_priv / tot_colegios * 100) if tot_colegios > 0 else 0
tot_academias = len(df_comp_filtered)

pob_total = sum(SISE_CAMPUSES[s]['pob_2025'] for s in sedes_seleccionadas if s in SISE_CAMPUSES)
alumnos_meta = sum(SISE_CAMPUSES[s]['alumnos_obj'] for s in sedes_seleccionadas if s in SISE_CAMPUSES)
mercado_asignado = sum(SISE_CAMPUSES[s]['mercado_asignado'] for s in sedes_seleccionadas if s in SISE_CAMPUSES)

with col1:
    st.markdown(f'<div class="kpi-card"><div class="kpi-title">Colegios Activos (5km)</div><div class="kpi-value">{tot_colegios:,}</div><div class="kpi-sub">🟢 {len(sedes_seleccionadas)} Sedes SISE</div></div>', unsafe_allow_html=True)
with col2:
    st.markdown(f'<div class="kpi-card"><div class="kpi-title">% Colegios Privados</div><div class="kpi-value">{pct_priv:.1f}%</div><div class="kpi-sub">🏫 {tot_priv:,} Priv. / {tot_pub:,} Púb.</div></div>', unsafe_allow_html=True)
with col3:
    st.markdown(f'<div class="kpi-card"><div class="kpi-title">Academia Competencia</div><div class="kpi-value">{tot_academias}</div><div class="kpi-sub">📙 En Radio de 5 km</div></div>', unsafe_allow_html=True)
with col4:
    st.markdown(f'<div class="kpi-card"><div class="kpi-title">Población Distrital 2025</div><div class="kpi-value">{pob_total:,.0f}</div><div class="kpi-sub">👥 Fuente INEI 2025</div></div>', unsafe_allow_html=True)
with col5:
    st.markdown(f'<div class="kpi-card"><div class="kpi-title">Mercado Preu Asignado</div><div class="kpi-value">{mercado_asignado:,.0f}</div><div class="kpi-sub">🎯 {alumnos_meta:,.0f} Meta SISE</div></div>', unsafe_allow_html=True)

st.markdown("<br>", unsafe_allow_html=True)

# =============================================================================
# BLOQUE 2: GRÁFICOS INTERACTIVOS — CON BARRAS COMPARATIVAS DUALES (ALUMNOS OBJETIVO VS MERCADO ASIGNADO)
# =============================================================================
st.subheader("📊 » Análisis de Potencial de Mercado y Competencia por Sede SISE")

t0, t1, t2, t3 = st.tabs([
    "🎯 Potencial & Embudo de Mercado",
    "🏛️ Colegios por Sede (Públicos vs Privados)", 
    "📙 Competencia por Marca", 
    "🎯 Matriz Océano Azul (Potencial vs Competencia)"
])

with t0:
    st.markdown("### » 1. Modelo Macro de Estimación de Mercado de Preuniversitario (Lima 2025 - 2026)")
    st.markdown("El mercado total se calcula multiplicando los postulantes proyectados al 2026 (+2.0% crecimiento anual) por el % de alcanzabilidad en Lima y el % de alumnos que contrata academia. El objetivo corporativo SISE es captar el **5.0% de Market Share**.")
    
    if not df_mercado_macro.empty:
        df_m_display = df_mercado_macro.copy()
        df_m_display['Postulantes 2025'] = df_m_display['Postulantes 2025'].map('{:,.0f}'.format)
        df_m_display['Ingresantes 2025'] = df_m_display['Ingresantes 2025'].map('{:,.0f}'.format)
        df_m_display['% No Ingresa (info)'] = df_m_display['% No Ingresa (info)'].map('{:.1%}'.format)
        df_m_display['Postulantes 2026 (proy.)'] = df_m_display['Postulantes 2026 (proy.)'].map('{:,.0f}'.format)
        df_m_display['% Alcanzable (Lima)'] = df_m_display['% Alcanzable (Lima)'].apply(lambda x: f"{x:.1%}" if pd.notnull(x) else "-")
        df_m_display['% Usa Academia'] = df_m_display['% Usa Academia'].apply(lambda x: f"{x:.1%}" if pd.notnull(x) else "-")
        df_m_display['Mercado de Preu'] = df_m_display['Mercado de Preu'].map('{:,.0f}'.format)
        df_m_display['% Market Share'] = df_m_display['% Market Share'].apply(lambda x: f"{x:.1%}" if pd.notnull(x) else "-")
        df_m_display['Alumnos Objetivo'] = df_m_display['Alumnos Objetivo'].map('{:,.0f}'.format)
        
        st.dataframe(df_m_display, use_container_width=True)

    st.markdown("### » 2. Comparativo de Alumnos Objetivo Meta SISE (5% Captación) vs Mercado Preu Asignado por Sede")
    
    col_f1, col_f2 = st.columns([1, 1])
    
    with col_f1:
        # EMBUDO DE CONVERSIÓN
        funnel_data = dict(
            number=[128806, 131383, 83721, 61002, 3050],
            stage=["Postulantes Totales '25", "Postulantes Proy. '26 (+2%)", "Alcanzables Residentes Lima", "Mercado Preu (Usa Academia)", "Alumnos Objetivo SISE (5% Share)"]
        )
        fig_funnel = px.funnel(funnel_data, x='number', y='stage', title="Embudo Metodológico de Estimación de Demanda", color_discrete_sequence=['#FF0E49'])
        fig_funnel.update_layout(height=450, font=dict(family="Barlow Condensed"))
        st.plotly_chart(fig_funnel, use_container_width=True)
        
    with col_f2:
        # GRÁFICO COMPARATIVO DUAL
        potencial_list = []
        for cname in sedes_seleccionadas:
            info_c = SISE_CAMPUSES[cname]
            potencial_list.append({
                "Sede SISE": cname.replace("SISE ", ""),
                "Alumnos Objetivo (Meta SISE 5%)": info_c['alumnos_obj'],
                "Mercado Preu Asignado Total": info_c['mercado_asignado']
            })
        df_pot = pd.DataFrame(potencial_list).sort_values(by="Alumnos Objetivo (Meta SISE 5%)", ascending=True)

        df_pot_grouped = pd.melt(
            df_pot, 
            id_vars=["Sede SISE"], 
            value_vars=["Alumnos Objetivo (Meta SISE 5%)", "Mercado Preu Asignado Total"],
            var_name="Métrica", 
            value_name="Cantidad Alumnos"
        )

        fig_pot = px.bar(
            df_pot_grouped, y="Sede SISE", x="Cantidad Alumnos", color="Métrica",
            title="Potencial por Sede: Alumnos Objetivo Meta SISE vs Mercado Preu Asignado",
            orientation="h", barmode="group",
            text="Cantidad Alumnos",
            color_discrete_map={
                "Alumnos Objetivo (Meta SISE 5%)": "#FF0E49",
                "Mercado Preu Asignado Total": "#0075B0"
            }
        )
        fig_pot.update_traces(texttemplate='%{text:,.0f}', textposition='outside')
        fig_pot.update_layout(xaxis_title="Cantidad de Alumnos", yaxis_title="", height=450, font=dict(family="Barlow Condensed"), legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1))
        st.plotly_chart(fig_pot, use_container_width=True)

with t1:
    summary_list = []
    for cname in sedes_seleccionadas:
        df_sub = df_sch_filtered[df_sch_filtered['sede_sise'] == cname] if not df_sch_filtered.empty else pd.DataFrame()
        summary_list.append({
            "Sede SISE": cname.replace("SISE ", ""),
            "Públicos": len(df_sub[df_sub['tipo_gestion'] == 'Público']) if not df_sub.empty else 0,
            "Privados": len(df_sub[df_sub['tipo_gestion'] == 'Privado']) if not df_sub.empty else 0,
            "Total Colegios": len(df_sub)
        })
    df_sum = pd.DataFrame(summary_list).sort_values(by="Total Colegios", ascending=True)

    fig_bar = px.bar(
        df_sum, y="Sede SISE", x=["Privados", "Públicos"],
        title="Distribución de Colegios Activos a 5 km (Privados vs Públicos)",
        orientation="h", barmode="stack",
        color_discrete_map={"Privados": "#0075B0", "Públicos": "#009860"}
    )
    fig_bar.update_layout(xaxis_title="Cantidad de Colegios Activos", yaxis_title="", height=420, font=dict(family="Barlow Condensed"))
    st.plotly_chart(fig_bar, use_container_width=True)

with t2:
    if not df_comp_filtered.empty:
        comp_summary = df_comp_filtered.groupby(['cadena']).size().reset_index(name='Cantidad de Locales')
        color_map = {
            "Aula 20": "#00a8ff", "Trilce": "#e67e22", "Pamer": "#bdc3c7",
            "César Vallejo y ADUNI": "#8e44ad", "CEPREPUCP": "#0c2461",
            "Pre San Marcos": "#800020", "CEPRE UNI": "#f9e79f"
        }
        fig_pie = px.pie(
            comp_summary, values='Cantidad de Locales', names='cadena',
            title='Participación por Cadena de Academias Competencia Seleccionadas',
            color='cadena', color_discrete_map=color_map, hole=0.4
        )
        fig_pie.update_layout(height=420, font=dict(family="Barlow Condensed"))
        st.plotly_chart(fig_pie, use_container_width=True)
    else:
        st.info("No hay marcas de competencia seleccionadas.")

with t3:
    matrix_list = []
    for cname in sedes_seleccionadas:
        df_s = df_sch_filtered[df_sch_filtered['sede_sise'] == cname] if not df_sch_filtered.empty else pd.DataFrame()
        cnt_comp = len(df_comp_filtered[df_comp_filtered['sedes_5km'].apply(lambda x: cname in x)]) if not df_comp_filtered.empty else 0
        matrix_list.append({
            "Sede": cname.replace("SISE ", ""),
            "Colegios Activos (5km)": len(df_s),
            "Competencia Academias (5km)": cnt_comp,
            "Alumnos Meta": SISE_CAMPUSES[cname]['alumnos_obj']
        })
    df_matrix = pd.DataFrame(matrix_list)

    if not df_matrix.empty:
        fig_scat = px.scatter(
            df_matrix, x="Colegios Activos (5km)", y="Competencia Academias (5km)",
            size="Alumnos Meta", text="Sede", color="Colegios Activos (5km)",
            title="Matriz de Oportunidad: Colegios Activos vs Competencia Presente",
            color_continuous_scale=["#0075B0", "#FF0E49"], size_max=40
        )
        fig_scat.update_traces(textposition='top center')
        fig_scat.update_layout(height=450, font=dict(family="Barlow Condensed"))
        st.plotly_chart(fig_scat, use_container_width=True)

# =============================================================================
# BLOQUE 3: MAPA ULTRARRÁPIDO ACELERADO POR CANVAS (60 FPS)
# =============================================================================
st.subheader("🗺️ » Mapa Georeferenciado en Tiempo Real (Acelerado por Canvas)")

if len(sedes_seleccionadas) == 1:
    s_key = sedes_seleccionadas[0]
    center_lat, center_lon = SISE_CAMPUSES[s_key]['lat'], SISE_CAMPUSES[s_key]['lon']
    zoom_level = 13
else:
    center_lat, center_lon = -12.08, -77.03
    zoom_level = 11

svg_sise_logo = 'data:image/svg+xml;utf8,<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 32 40" width="32" height="40"><path fill="%23FF0E49" stroke="%23231F20" stroke-width="2.5" d="M16 2C8.27 2 2 8.27 2 16c0 11.2 14 22 14 22s14-10.8 14-22c0-7.73-6.27-14-14-14z"/><circle cx="16" cy="16" r="8" fill="%23FFFFFF"/><text x="16" y="20.5" font-family="Arial,sans-serif" font-size="11" font-weight="900" fill="%23FF0E49" text-anchor="middle">S</text></svg>'

campuses_json = json.dumps(
    {k: v for k, v in SISE_CAMPUSES.items() if k in sedes_seleccionadas},
    ensure_ascii=False
)
schools_json = json.dumps(df_sch_filtered.to_dict(orient='records'), ensure_ascii=False) if not df_sch_filtered.empty else "[]"
academias_json = json.dumps(df_comp_filtered.to_dict(orient='records'), ensure_ascii=False) if not df_comp_filtered.empty else "[]"

fast_canvas_map_html = f"""<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="UTF-8">
    <link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css" />
    <script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"></script>
    <style>
        html, body {{ margin:0; padding:0; width:100%; height:100%; overflow:hidden; font-family:'Segoe UI',Arial,sans-serif; }}
        #map {{ width:100%; height:100vh; }}
        .review-box {{ background:#fff9e6; border-left:3px solid #FF0E49; padding:6px; margin-top:5px; font-size:11px; color:#231F20; font-style:italic; border-radius:3px; }}
        .review-item {{ margin-bottom:3px; padding-bottom:2px; border-bottom:1px dashed #FF0E49; }}
    </style>
</head>
<body>
<div id="map"></div>

<script>
    const map = L.map('map', {{ preferCanvas: true }}).setView([{center_lat}, {center_lon}], {zoom_level});
    L.tileLayer('https://{{s}}.tile.openstreetmap.org/{{z}}/{{x}}/{{y}}.png', {{ maxZoom: 18 }}).addTo(map);

    function buildPinSvg(fillColor, strokeColor) {{
        return `<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 32 40" width="26" height="34">
            <path fill="${{fillColor}}" stroke="${{strokeColor}}" stroke-width="2.5" d="M16 2C8.27 2 2 8.27 2 16c0 11.2 14 22 14 22s14-10.8 14-22c0-7.73-6.27-14-14-14z"/>
            <path fill="${{fillColor === '#ffffff' || fillColor === '#f9e79f' ? strokeColor : '#ffffff'}}" d="M16 8l-8 4.5 8 4.5 8-4.5zm-5.5 7.5v3.5l5.5 3 5.5-3v-3.5l-5.5 3z"/>
        </svg>`;
    }}

    function getAcademiaIcon(cadena) {{
        let fill = '#e67e22';
        let stroke = '#ffffff';
        let cUpper = (cadena || '').toUpperCase();
        if (cUpper.includes('AULA 20')) fill = '#00a8ff';
        else if (cUpper.includes('TRILCE')) fill = '#e67e22';
        else if (cUpper.includes('PAMER')) {{ fill = '#ffffff'; stroke = '#231F20'; }}
        else if (cUpper.includes('VALLEJO') || cUpper.includes('ADUNI')) fill = '#8e44ad';
        else if (cUpper.includes('PUCP')) fill = '#0c2461';
        else if (cUpper.includes('SAN MARCOS')) fill = '#800020';
        else if (cUpper.includes('UNI')) {{ fill = '#f9e79f'; stroke = '#7d6608'; }}

        return L.icon({{
            iconUrl: 'data:image/svg+xml;utf8,' + encodeURIComponent(buildPinSvg(fill, stroke)),
            iconSize: [26, 34], iconAnchor: [13, 34], popupAnchor: [0, -32]
        }});
    }}

    const siseLogoIcon = L.icon({{
        iconUrl: '{svg_sise_logo}',
        iconSize: [30, 38], iconAnchor: [15, 38], popupAnchor: [0, -36]
    }});

    const campuses = {campuses_json};
    const schools = {schools_json};
    const academias = {academias_json};

    for (let cname in campuses) {{
        let c = campuses[cname];
        L.marker([c.lat, c.lon], {{icon: siseLogoIcon}}).addTo(map)
         .bindPopup(`<b>SEDE SISE: ${{cname}}</b><br>${{c.direccion}}<br><i>${{c.distrito}}</i>`);
        L.circle([c.lat, c.lon], {{ color: '#FF0E49', fillColor: '#FF0E49', fillOpacity: 0.08, radius: {RADIO_KM * 1000} }}).addTo(map);
    }}

    academias.forEach(a => {{
        let icon = getAcademiaIcon(a.cadena);
        let marker = L.marker([a.latitud, a.longitud], {{icon: icon}}).addTo(map);
        
        let reviewHtml = '';
        if (a.resumen_resenas) {{
            let parts = a.resumen_resenas.split('|').map(p => p.trim()).filter(p => p.length > 0);
            reviewHtml = '<div class="review-box"><b>💬 Feedback Reseñas:</b><br>' + 
                         parts.map(p => `<div class="review-item">"💬 ${{p}}"</div>`).join('') + 
                         '</div>';
        }}
        
        marker.bindPopup(`
            <b style="font-size:13px; color:#FF0E49;">${{a.nombre}}</b><br>
            <b>Cadena:</b> ${{a.cadena}}<br>
            <b>Dirección:</b> ${{a.direccion}}<br>
            <b>Sede SISE cercana:</b> ${{a.sede_sise_cercana}} (${{a.distancia_sise_km}} km)<br>
            <b>Calificación:</b> ⭐ ${{a.calificacion}} (${{a.total_resenas}} reseñas)<br>
            ${{reviewHtml}}
            <a href="${{a.enlace_gmaps}}" target="_blank" style="display:inline-block; margin-top:4px; color:#FF0E49; font-weight:bold;">Ver en Google Maps</a>
        `);
    }});

    schools.forEach(s => {{
        let dotColor = s.tipo_gestion === 'Público' ? '#009860' : '#0075B0';
        let circleMarker = L.circleMarker([s.latitud, s.longitud], {{
            radius: 3.5,
            fillColor: dotColor,
            color: dotColor,
            weight: 0,
            fillOpacity: 0.85
        }}).addTo(map);

        circleMarker.bindPopup(`
            <b>${{s.nombre_colegio}}</b><br>
            <b>Gestión:</b> ${{s.tipo_gestion}} (${{s.gestion}})<br>
            <b>Sede SISE:</b> ${{s.sede_sise}} (${{s.distancia_km}} km)<br>
            <b>Dirección:</b> ${{s.direccion}}<br>
            <b>Distrito:</b> ${{s.distrito}} (${{s.provincia}})
        `);
    }});
</script>
</body>
</html>"""

components.html(fast_canvas_map_html, height=540, scrolling=False)

# =============================================================================
# BLOQUE 4: TABLAS Y ACORDEÓN ERGONÓMICO COMPACTO (EXPANDERS ANIDADOS)
# =============================================================================
st.subheader("📋 » Análisis de Competencia & Modelo de Demanda Distrital")

tab_acordeon, tab_demanda, tab_tabla_comp, tab_col, tab_resumen = st.tabs([
    "💬 ESTRUCTURA COMPACTA DE RESEÑAS POR MARCA Y SEDE", 
    "📈 MODELO DE DEMANDA Y MERCADO ASIGNADO",
    "📙 Academias Tabla Completa", 
    "🏫 Colegios Activos MINEDU", 
    "📊 Tabla Resumen Ejecutivo"
])

with tab_acordeon:
    st.markdown("### » Árbol Desplegable Compacto de Reseñas por Marca y Sede")
    
    if not df_comp_filtered.empty:
        cadenas_presentes = sorted(list(df_comp_filtered['cadena'].unique()))
        
        for cad in cadenas_presentes:
            df_cad = df_comp_filtered[df_comp_filtered['cadena'] == cad]
            cant_locales = len(df_cad)
            
            with st.expander(f"🏫 MARCA: {cad.upper()} ({cant_locales} locales en Lima)", expanded=False):
                summary_text = BRAND_EXECUTIVE_SUMMARIES.get(cad, f"📌 **{cad}**: Cadena de preparación preuniversitaria.")
                st.markdown(f'<div class="brand-summary-card">{summary_text}</div>', unsafe_allow_html=True)
                
                for idx, row in df_cad.iterrows():
                    sede_title = f"📍 {row['nombre']} — ⭐ {row['calificacion']} ({row['total_resenas']} reseñas)"
                    
                    with st.expander(sede_title, expanded=False):
                        st.markdown(f"""
                        <div class="sede-meta-line">
                            <b>🏢 Dirección:</b> {row['direccion']} &nbsp;|&nbsp; 
                            <b>🎯 SISE Cercana:</b> {row['sede_sise_cercana']} ({row['distancia_sise_km']} km)
                        </div>
                        """, unsafe_allow_html=True)
                        
                        raw_res = str(row['resumen_resenas'] or '').strip()
                        if raw_res and raw_res != 'Sin reseñas destacadas':
                            parts = [p.strip() for p in raw_res.split('|') if p.strip()]
                            for p_idx, p_text in enumerate(parts, 1):
                                st.markdown(f'<div class="review-quote-box"><b>💬 Reseña #{p_idx}:</b> "{p_text}"</div>', unsafe_allow_html=True)
                        else:
                            st.info("Sin reseñas cualitativas destacadas para este local.")
                        
                        if row['enlace_gmaps']:
                            st.markdown(f'<a href="{row["enlace_gmaps"]}" target="_blank" style="font-size: 11px; font-weight: bold; color: #FF0E49;">🔗 Abrir Local en Google Maps</a>', unsafe_allow_html=True)
    else:
        st.info("No hay marcas de competencia que coincidan con los filtros seleccionados.")

with tab_demanda:
    st.markdown("### » Detalle del Modelo de Demanda (Población Distrital y Reparto de Postulantes)")
    if not df_demanda_sede.empty:
        st.dataframe(df_demanda_sede, use_container_width=True)
        csv_dem = df_demanda_sede.to_csv(index=False).encode('utf-8-sig')
        st.download_button("📥 Descargar Modelo de Demanda (CSV)", csv_dem, "Demanda_Preu_SISE_Sedes.csv", "text/csv")
    else:
        st.info("No se encontró el archivo de demanda.")

with tab_tabla_comp:
    if not df_comp_filtered.empty:
        st.dataframe(df_comp_filtered[['cadena', 'nombre', 'sede_sise_cercana', 'distancia_sise_km', 'calificacion', 'total_resenas', 'resumen_resenas', 'direccion', 'enlace_gmaps']], use_container_width=True)
        csv_comp = df_comp_filtered.to_csv(index=False).encode('utf-8-sig')
        st.download_button("📥 Descargar Competencia Filtrada con Reseñas (CSV)", csv_comp, "Competencia_SISE_Con_Resenas.csv", "text/csv")

with tab_col:
    if not df_sch_filtered.empty:
        st.dataframe(df_sch_filtered[['sede_sise', 'distancia_km', 'nombre_colegio', 'tipo_gestion', 'gestion', 'provincia', 'distrito', 'direccion', 'cod_mod']], use_container_width=True)
        csv_col = df_sch_filtered.to_csv(index=False).encode('utf-8-sig')
        st.download_button("📥 Descargar Colegios Filtrados (CSV)", csv_col, "Colegios_SISE_Filtrados.csv", "text/csv")
    else:
        st.info("No hay colegios filtrados.")

with tab_resumen:
    st.dataframe(df_matrix, use_container_width=True)

# FOOTER OFICIAL SISE
st.markdown("""
<div class="sise-footer">
    SISE INSTITUTO SUPERIOR · GRUPO EDUCAD · #HAZLABIEN<br>
    <span style="font-weight:400; opacity:0.8;">Call Center 625 5656 · WhatsApp 920 141 309 · www.sise.edu.pe</span>
</div>
""", unsafe_allow_html=True)
